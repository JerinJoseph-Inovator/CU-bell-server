#!/usr/bin/env python3
import RPi.GPIO as GPIO
import time
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os
relay_pin=12
GPIO.setmode(GPIO.BCM)
GPIO.setup(relay_pin,GPIO.OUT)
# Function to simulate ringing the bell
def ring_bell(duration):
    GPIO.output(relay_pin,GPIO.HIGH)
    print(f"Bell ringing for {duration} seconds", flush=True)
    time.sleep(duration)
    print("Bell stopped", flush=True)
    GPIO.output(relay_pin,GPIO.LOW)

saturday_timings = [
    "08:55:00", "09:00:00", "09:55:00", "10:00:00", "10:55:00", "11:00:00",
    "11:55:00", "12:00:00", "12:55:00"
]
weekday_timings = [
    "08:55:00", "09:00:00", "09:55:00", "10:00:00", "10:55:00", "11:00:00",
    "11:55:00", "12:00:00", "12:55:00", "13:00:00", "13:55:00", "14:00:00",
    "14:55:00", "15:00:00", "15:55:00"
]
weekdays=[
    "Monday","Tuesday","Wednesday","Thursday","Friday"
]

# Function to process the latest events file
def process_latest_events(input_file):
    date_events = {}
    immediate_ring = False  # Flag for immediate bell ring
    specific_time_event = None  # To track specific time events

    with open(input_file, 'r') as infile:
        for line in infile:
            line = line.strip()
            if line:
                parts = line.split(',')
                event_type = int(parts[0])

                if event_type == 0:
                    # Holiday event
                    date_str = parts[1]
                    date_events[date_str] = {'event': (event_type, None, None)}
                elif event_type in (1, 2):
                    # MID SEM or END SEM
                    slot = int(parts[1])
                    date_str = parts[2]
                    event_time = parts[3] if len(parts) > 3 else None

                    # Add event if not a holiday
                    if date_str not in date_events or date_events[date_str].get('event', (None,))[0] != 0:
                        if date_str not in date_events:
                            date_events[date_str] = {}
                        date_events[date_str][slot] = event_time
                elif event_type == 3:
                    # Specific time bell event
                    specific_time_event = parts[1]
                    immediate_ring = True

    return date_events, immediate_ring, specific_time_event

# Offsets for MID SEM and END SEM bell timings
midsem_offsets = {
    1: [timedelta(minutes=57), timedelta(minutes=30), timedelta(minutes=43), timedelta(hours=2, minutes=15), timedelta(hours=2, minutes=30)],
    2: [timedelta(minutes=13, seconds=45), timedelta(minutes=35), timedelta(minutes=48), timedelta(hours=2, minutes=20), timedelta(hours=2, minutes=35)],
    3: [timedelta(minutes=13), timedelta(minutes=40), timedelta(minutes=53), timedelta(hours=2, minutes=25), timedelta(hours=2, minutes=40)]
}

endsem_offsets = {
    1: [timedelta(minutes=15), timedelta(minutes=32), timedelta(minutes=30), timedelta(hours=3), timedelta(hours=3, minutes=30)],
    2: [timedelta(minutes=20), timedelta(minutes=37), timedelta(minutes=35), timedelta(hours=3, minutes=5), timedelta(hours=3, minutes=35)],
    3: [timedelta(minutes=46), timedelta(minutes=42), timedelta(minutes=40), timedelta(hours=3, minutes=10), timedelta(hours=3, minutes=40)]
}

# Function to calculate bell times for MID SEM or END SEM
def calculate_bell_times(base_time_str, offsets):
    base_time = datetime.strptime(base_time_str, "%H:%M:%S")
    return [(base_time + offset).strftime("%H:%M:%S") for offset in offsets]

# Function to update the Excel file
def update_excel_file(date_str, event_data):
    excel_file_path = "bell_events.xlsx"

    # Load existing workbook or create a new one
    if os.path.exists(excel_file_path):
        workbook = load_workbook(excel_file_path)
    else:
        workbook = Workbook()

    # Select the active sheet
    sheet = workbook.active
    sheet.title = "Bell Events"

    # Remove previous entries for the same date
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value == date_str:
            row_index = row[0].row
            sheet.delete_rows(row_index)

    # Add new entries for the date
    for event_type, slot, timings in event_data:
        sheet.append([date_str, event_type, slot, ", ".join(timings)])

    workbook.save(excel_file_path)
    print(f"Excel file updated for {date_str}.")

# Main loop
def main_loop():
    print("Starting the main loop...")

    while True:
        now = datetime.now()
        current_time_str = now.strftime("%H:%M:%S")
        current_date = now.strftime("%d-%m-%Y")
        current_weekday = now.strftime("%A")

        print(f"Checking events at {current_time_str} on {current_date} ({current_weekday})")

        # Process events from the input file
        date_events, immediate_ring, specific_time_event = process_latest_events("/home/pi/Desktop/server/input.txt")

        # Print events for the current date
        if current_date in date_events:
            print(f"Events for {current_date}:")
            event_info = date_events[current_date]
            for key, value in event_info.items():
                if key == 'event':
                    print(f"Type: {value[0]}, Slot: {value[1]}, Time: {value[2]}")
                else:
                    print(f"Slot: {key}, Time: {value}")
        else:
            print(f"No events scheduled for {current_date}.")

        # Immediate bell ring
        if immediate_ring:
            print(f"Immediate bell ring triggered at {current_time_str}")
            ring_bell(5)
        
        if current_weekday == "Sunday":
            time.sleep(60)
            continue

        if current_weekday == "Saturday":
            if current_time_str in saturday_timings:
                ring_bell(5)
        
        if current_weekday in weekdays:
            if current_time_str in weekday_timings:
                ring_bell(5)
        # Process today's events
        excel_updates = []

        if current_date in date_events:
            event_info = date_events[current_date]

            # Handle MID SEM and END SEM
            if isinstance(event_info, dict) and 'event' not in event_info:
                for slot, time_str in event_info.items():
                    if time_str:
                        offsets = midsem_offsets if 1 in event_info else endsem_offsets
                        result_times = calculate_bell_times(time_str, offsets[slot])
                        excel_updates.append((1, slot, result_times))

                        if current_time_str in result_times:
                            ring_bell(8)
            elif 'event' in event_info:
                event_type, slot, event_time = event_info['event']
                if event_type == 0:
                    print("Holiday today")
                    excel_updates.append((0, None, []))

        if excel_updates:
            update_excel_file(current_date, excel_updates)

        time.sleep(1)

# Run the main loop
main_loop()
